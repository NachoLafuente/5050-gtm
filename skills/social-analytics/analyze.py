#!/usr/bin/env python3
"""
social-analytics: turn a LinkedIn Creator analytics export into a
"what works / what doesn't" report.

Reads the AggregateAnalytics_*.xlsx that LinkedIn gives you (Analytics tab ->
Export) and, optionally, the Shares_*.csv from your full data archive so each
top post is joined to its full text. Tags every post by topic / hook / day /
length, computes engagement rate, ranks winners and losers, and writes a styled
Excel workbook plus a tagged CSV.

No API keys. No dashboards. One shot, local files in, report + xlsx out.

Usage:
    python analyze.py --analytics "AggregateAnalytics_Name_2025-..-...xlsx"
    python analyze.py --analytics ana.xlsx --archive Shares_123.csv
    python analyze.py --analytics ana.xlsx --archive ~/Desktop/Complete_LinkedInDataExport_folder
    python analyze.py --analytics ana.xlsx --out ./out --top 50

The --archive value can point at the Shares_*.csv directly, or at the unzipped
"Complete_LinkedInDataExport" folder (the script will find Shares_*.csv inside).
"""

import argparse
import csv
import glob
import json
import math
import os
import re
import statistics as st
import sys
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing openpyxl. Run: pip install openpyxl>=3.1")


# ----------------------------------------------------------------------------
# Taxonomy. Edit these two dicts to match your niche — defaults are tuned for a
# B2B / GTM / CRM personal brand. Keys are labels, values are regexes matched
# (case-insensitive) against the post text.
# ----------------------------------------------------------------------------
TOPICS = {
    "Product/Tool": r"\b(product|tool|tools|built|cli|app|feature|launch|drop|shipped)\b",
    "AI agents": r"\bagent|\bllm\b|\bgpt\b|\bclaude\b|\bai\b",
    "Automation": r"\bautomat|workflow|\bn8n\b|zapier|integrat",
    "CRM": r"\bcrm\b|salesforce|hubspot|pipedrive|attio|close\.com",
    "GTM/Sales": r"\bgtm\b|go.to.market|outbound|pipeline|prospect|cold (email|call)",
    "Migration/Data": r"migrat|dedup|duplicat|enrich|data quality|import",
    "Hiring/Team": r"\bhiring\b|\bteam\b|\bfounder|\bclient",
    "Opinion/Story": r"\bi \b|\bmy \b|honest|lesson|learned|mistake|unpopular|hot take",
}

# Hook = how the FIRST line opens. A post can match several; all are credited.
HOOKS = {
    "Question": lambda first: "?" in first[:120],
    "Number/Stat": lambda first: bool(re.search(r"^\s*\d|\b\d+[\+xX%]|\b\d{2,}\b", first[:80])),
    "Personal-I": lambda first: bool(re.match(r"^(i |my |last |when i|how i|why i)", first.lower())),
    "Contrarian": lambda first: bool(
        re.search(r"everyone|nobody|most |stop |unpopular|hot take|wrong|myth|never|don.t", first.lower())
    ),
    "How-to": lambda first: bool(re.search(r"how to|here.s how|the \w+ way|steps?|guide|playbook", first.lower())),
}

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(color="FFFFFF", bold=True)


# ----------------------------------------------------------------------------
# Parsing
# ----------------------------------------------------------------------------
def parse_analytics(path):
    """Return dict of the five LinkedIn analytics sheets, defensively."""
    wb = load_workbook(path, data_only=True)
    out = {"summary": {}, "daily": [], "top_posts": [], "followers": [], "demographics": []}

    def rows(name):
        if name not in wb.sheetnames:
            return []
        return list(wb[name].iter_rows(values_only=True))

    # DISCOVERY: label/value pairs
    for r in rows("DISCOVERY"):
        if r and r[0] and r[1] is not None and not str(r[0]).startswith("Overall"):
            out["summary"][str(r[0])] = r[1]

    # ENGAGEMENT: Date, Impressions, Engagements
    for r in rows("ENGAGEMENT")[1:]:
        if r and r[0]:
            out["daily"].append({"date": r[0], "impressions": r[1], "engagements": r[2]})

    # FOLLOWERS: total on row 0, then Date/New followers table
    fr = rows("FOLLOWERS")
    if fr:
        out["summary"]["Total followers"] = fr[0][1] if fr[0] and len(fr[0]) > 1 else None
        for r in fr:
            if r and r[0] and str(r[0]) not in ("Date",) and not str(r[0]).startswith("Total"):
                try:
                    int(r[1])
                    out["followers"].append({"date": r[0], "new_followers": r[1]})
                except (TypeError, ValueError):
                    pass

    # DEMOGRAPHICS: Category, Value, Percentage
    for r in rows("DEMOGRAPHICS")[1:]:
        if r and r[0]:
            out["demographics"].append({"category": r[0], "value": r[1], "percentage": r[2]})

    # TOP POSTS: two side-by-side ranked lists (eng | imp). Find the header row.
    tp = rows("TOP POSTS")
    hdr_idx = next((i for i, r in enumerate(tp) if r and str(r[0]).strip() == "Post URL"), None)
    eng, imp, date = {}, {}, {}
    if hdr_idx is not None:
        for r in tp[hdr_idx + 1:]:
            if r and r[0]:
                eng[r[0]] = r[2]
                date[r[0]] = r[1]
            if r and len(r) > 4 and r[4]:
                imp[r[4]] = r[6]
                date.setdefault(r[4], r[5])
    for u in set(eng) | set(imp):
        out["top_posts"].append(
            {"url": u, "date": date.get(u, ""), "engagements": eng.get(u), "impressions": imp.get(u)}
        )
    return out


def find_shares_csv(archive):
    if not archive:
        return None
    if os.path.isfile(archive):
        return archive
    if os.path.isdir(archive):
        hits = glob.glob(os.path.join(archive, "**", "Shares_*.csv"), recursive=True)
        return hits[0] if hits else None
    return None


def slug_from_url(u):
    m = re.search(r"_(.+?)-(?:ugcPost|activity|share)-", u)
    return m.group(1) if m else ""


def norm(s):
    s = re.sub(r"https?://\S+", " ", s.lower())
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


def join_text(top_posts, shares_csv):
    """Match analytics URLs to archive post text by the text-slug in the URL."""
    if not shares_csv:
        for p in top_posts:
            p["text"], p["media"] = "", ""
        return 0
    shares = []
    with open(shares_csv, newline="", encoding="utf-8", errors="replace") as f:
        for s in csv.DictReader(f):
            t = (s.get("ShareCommentary") or "").strip()
            shares.append((norm(t)[:120], t, s.get("MediaUrl", "")))
    matched = 0
    for p in top_posts:
        slug = slug_from_url(p["url"])
        best, best_len = None, 0
        if slug:
            for sn, txt, media in shares:
                common = 0
                for a, b in zip(slug, sn):
                    if a == b:
                        common += 1
                    else:
                        break
                if common > best_len and common >= 15:
                    best_len, best = common, (txt, media)
        p["text"], p["media"] = (best[0], best[1]) if best else ("", "")
        if best:
            matched += 1
    return matched


# ----------------------------------------------------------------------------
# Enrichment + stats
# ----------------------------------------------------------------------------
def tag_topics(text):
    tl = text.lower()
    out = [k for k, rx in TOPICS.items() if re.search(rx, tl)]
    return out or ["Other"]


def tag_hooks(text):
    first = text.strip().split("\n")[0] if text.strip() else ""
    out = [k for k, fn in HOOKS.items() if fn(first)]
    return out or ["Statement"]


def parse_date(d):
    if isinstance(d, datetime):
        return d
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(str(d), fmt)
        except (ValueError, TypeError):
            continue
    return None


def enrich(top_posts):
    P = []
    for p in top_posts:
        try:
            imp = float(p["impressions"])
            eng = float(p["engagements"])
        except (TypeError, ValueError):
            continue
        d = parse_date(p["date"])
        text = p.get("text", "") or ""
        P.append(
            {
                **p,
                "imp": imp,
                "eng": eng,
                "er": (eng / imp * 100) if imp else 0,
                "dt": d,
                "len": len(text),
                "topics": tag_topics(text),
                "hooks": tag_hooks(text),
                "day": DAYS[d.weekday()] if d else "?",
            }
        )
    return P


def pearson(xs, ys):
    n = len(xs)
    if n < 2:
        return 0.0
    mx, my = st.mean(xs), st.mean(ys)
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    den = math.sqrt(sum((x - mx) ** 2 for x in xs) * sum((y - my) ** 2 for y in ys))
    return num / den if den else 0.0


def length_bucket(n):
    return "<300" if n < 300 else "300-800" if n < 800 else "800-1500" if n < 1500 else "1500+"


def agg(P, key_fn):
    g = defaultdict(list)
    for p in P:
        for k in key_fn(p):
            g[k].append(p)
    rows = []
    for k, ps in g.items():
        rows.append(
            {
                "key": k,
                "n": len(ps),
                "avg_imp": st.mean([x["imp"] for x in ps]),
                "avg_eng": st.mean([x["eng"] for x in ps]),
                "avg_er": st.mean([x["er"] for x in ps]),
            }
        )
    return sorted(rows, key=lambda r: -r["avg_eng"])


# ----------------------------------------------------------------------------
# Report (stdout)
# ----------------------------------------------------------------------------
def shorten(t, n=70):
    return (t.strip().replace("\n", " ")[:n]) if t else "(no text matched)"


def print_report(data, P, matched):
    s = data["summary"]
    imps = [p["imp"] for p in P]
    engs = [p["eng"] for p in P]
    ers = [p["er"] for p in P]
    L = []
    p = L.append

    p("=" * 70)
    p("LINKEDIN PERFORMANCE — what works / what doesn't")
    p("=" * 70)
    if s:
        p("\n## Summary")
        for k, v in s.items():
            p(f"  {k}: {v}")
    if not P:
        p("\nNo posts with both impressions AND engagements found. "
          "Check the TOP POSTS sheet in the analytics export.")
        print("\n".join(L))
        return

    p(f"\n## Dataset: {len(P)} posts (LinkedIn caps the analytics export at top ~50)")
    p(f"  Impressions  median {st.median(imps):.0f} | mean {st.mean(imps):.0f} | max {max(imps):.0f}")
    p(f"  Engagements  median {st.median(engs):.0f} | mean {st.mean(engs):.1f} | max {max(engs):.0f}")
    p(f"  Eng rate %%   median {st.median(ers):.2f} | mean {st.mean(ers):.2f} | max {max(ers):.2f}")
    p(f"  Text-matched to archive: {matched}/{len(P)}")

    p("\n  !! Engagement rate is inversely tied to reach: big posts dilute ER.")
    p("     Judge big posts by absolute engagement, small posts by ER.")
    p("     This is your TOP tier only (survivorship bias) — not what flops look like.")

    def block(title, rows):
        p(f"\n## By {title}  (avg imp / avg eng / avg ER% / n)")
        for r in rows:
            p(f"  {str(r['key']):16s} imp {r['avg_imp']:7.0f} | eng {r['avg_eng']:5.1f} "
              f"| ER {r['avg_er']:5.2f}% | n={r['n']}")

    block("TOPIC", agg(P, lambda x: x["topics"]))
    block("HOOK STYLE", agg(P, lambda x: x["hooks"]))
    block("DAY OF WEEK", agg(P, lambda x: [x["day"]]))
    block("LENGTH (chars)", agg(P, lambda x: [length_bucket(x["len"])]))

    p("\n## Correlations (Pearson)")
    p(f"  length vs engagements:      {pearson([x['len'] for x in P], engs):+.2f}")
    p(f"  impressions vs engagements: {pearson(imps, engs):+.2f}")

    p("\n## Top 8 by engagements")
    for x in sorted(P, key=lambda z: -z["eng"])[:8]:
        p(f"  eng {x['eng']:3.0f} | imp {x['imp']:6.0f} | ER {x['er']:4.1f}% | {x['day']} | {shorten(x.get('text',''))!r}")
    p("\n## Top 5 by eng rate (min 500 impressions)")
    for x in sorted([z for z in P if z["imp"] >= 500], key=lambda z: -z["er"])[:5]:
        p(f"  ER {x['er']:4.1f}% | eng {x['eng']:3.0f} | imp {x['imp']:6.0f} | {shorten(x.get('text',''))!r}")
    p("\n## Bottom 5 by engagements (your weakest top-tier posts)")
    for x in sorted(P, key=lambda z: z["eng"])[:5]:
        p(f"  eng {x['eng']:3.0f} | imp {x['imp']:6.0f} | {shorten(x.get('text',''))!r}")

    print("\n".join(L))


# ----------------------------------------------------------------------------
# Outputs (xlsx + csv)
# ----------------------------------------------------------------------------
def _sheet(wb, title, headers, rows, widths, wrapcol=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = HDR_FILL
        ws.cell(1, c).font = HDR_FONT
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    if wrapcol:
        for row in ws.iter_rows(min_row=2, min_col=wrapcol, max_col=wrapcol):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def write_xlsx(data, P, out_dir):
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.cell(1, 1).fill = HDR_FILL
    ws.cell(1, 1).font = HDR_FONT
    ws.cell(1, 2).fill = HDR_FILL
    ws.cell(1, 2).font = HDR_FONT
    for k, v in data["summary"].items():
        ws.append([k, v])
    if P:
        ws.append(["Posts analyzed", len(P)])
        ws.append(["Median engagement rate", f"{st.median([p['er'] for p in P]):.2f}%"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    # Post Performance
    pr = [
        [
            p["date"],
            p["imp"],
            p["eng"],
            f"{p['er']:.2f}%",
            ", ".join(p["topics"]),
            ", ".join(p["hooks"]),
            p["day"],
            p.get("text", ""),
            p["url"],
        ]
        for p in sorted(P, key=lambda z: -z["imp"])
    ]
    _sheet(
        wb,
        "Post Performance",
        ["Date", "Impressions", "Engagements", "Eng Rate", "Topics", "Hooks", "Day", "Post Text", "Post URL"],
        pr,
        [14, 13, 13, 11, 22, 20, 7, 80, 46],
        wrapcol=8,
    )

    # Daily Engagement
    _sheet(
        wb,
        "Daily Engagement",
        ["Date", "Impressions", "Engagements"],
        [[d["date"], d["impressions"], d["engagements"]] for d in data["daily"]],
        [16, 14, 14],
    )
    # Followers
    _sheet(
        wb,
        "Followers Daily",
        ["Date", "New Followers"],
        [[f["date"], f["new_followers"]] for f in data["followers"]],
        [16, 16],
    )
    # Demographics
    _sheet(
        wb,
        "Demographics",
        ["Category", "Value", "Percentage"],
        [[d["category"], d["value"], d["percentage"]] for d in data["demographics"]],
        [18, 40, 14],
    )

    path = os.path.join(out_dir, "linkedin_analysis.xlsx")
    wb.save(path)
    return path


def write_csv(P, out_dir):
    path = os.path.join(out_dir, "linkedin_posts_tagged.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Impressions", "Engagements", "EngRate%", "Topics", "Hooks", "Day", "Length", "PostText", "PostURL"])
        for p in sorted(P, key=lambda z: -z["imp"]):
            w.writerow(
                [
                    p["date"],
                    int(p["imp"]),
                    int(p["eng"]),
                    f"{p['er']:.2f}",
                    "|".join(p["topics"]),
                    "|".join(p["hooks"]),
                    p["day"],
                    p["len"],
                    (p.get("text", "") or "").replace("\n", " "),
                    p["url"],
                ]
            )
    return path


# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="LinkedIn analytics -> what-works report")
    ap.add_argument("--analytics", required=True, help="AggregateAnalytics_*.xlsx from LinkedIn")
    ap.add_argument("--archive", help="Shares_*.csv or the unzipped data-export folder (joins post text)")
    ap.add_argument("--out", default="./out", help="Output directory (default ./out)")
    ap.add_argument("--top", type=int, default=0, help="Limit to top N posts by impressions (0 = all)")
    ap.add_argument("--no-files", action="store_true", help="Print report only, skip xlsx/csv")
    args = ap.parse_args()

    if not os.path.isfile(args.analytics):
        sys.exit(f"Analytics file not found: {args.analytics}")

    data = parse_analytics(args.analytics)
    shares_csv = find_shares_csv(args.archive)
    matched = join_text(data["top_posts"], shares_csv)
    P = enrich(data["top_posts"])
    P.sort(key=lambda z: -z["imp"])
    if args.top:
        P = P[: args.top]

    print_report(data, P, matched)

    if not args.no_files and P:
        os.makedirs(args.out, exist_ok=True)
        xlsx = write_xlsx(data, P, args.out)
        csvp = write_csv(P, args.out)
        json.dump(P, open(os.path.join(args.out, "linkedin_posts.json"), "w"), default=str)
        print(f"\nWrote:\n  {xlsx}\n  {csvp}\n  {os.path.join(args.out, 'linkedin_posts.json')}")


if __name__ == "__main__":
    main()

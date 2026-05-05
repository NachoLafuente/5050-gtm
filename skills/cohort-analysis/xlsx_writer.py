"""Write the P9-styled cohort workbook with conditional formatting.

Layout mirrors the canonical Point Nine Cohort Analysis Template:
https://docs.google.com/spreadsheets/d/1WfI9EIQPiIccw3JkJfXgW6HqZU68BPsZcnSwnho7qgg/

One sheet, three sections stacked vertically:
  1. Customer Churn      (5 sub-tables)
  2. MRR Churn           (5 sub-tables)
  3. CAC Payback         (1 sub-table + 'profitable since' column)

Conditional formatting:
  - Retention %: red (low) → yellow → green (high)
  - Churn %:     green (low) → yellow → red (high), with green at the
                 negative end (expansion = good)
  - CAC payback: green if cumulative gross profit ≥ CAC
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="222222")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)


def _set_header(ws, row, col_start, col_end, label):
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start, end_row=row, end_column=col_end
        )


def _write_table(ws, top_row, left_col, title, columns, rows, data_fmt, color_rule):
    """Write a sub-table block: title row, header row, then data rows.

    columns: list of column labels (lifetime months or calendar months)
    rows:    list of (cohort_label, base_value, {col_idx: value})
    Returns: the bottom row of the data block (inclusive)
    """
    _set_header(ws, top_row, left_col, left_col + 1 + len(columns), title)

    # Header row
    hdr_row = top_row + 1
    ws.cell(row=hdr_row, column=left_col, value="Cohort").font = HEADER_FONT
    ws.cell(row=hdr_row, column=left_col).fill = HEADER_FILL
    ws.cell(row=hdr_row, column=left_col + 1, value="Base").font = HEADER_FONT
    ws.cell(row=hdr_row, column=left_col + 1).fill = HEADER_FILL
    for j, col in enumerate(columns):
        c = ws.cell(row=hdr_row, column=left_col + 2 + j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Data rows
    data_start = hdr_row + 1
    for i, (cohort, base, vals) in enumerate(rows):
        r = data_start + i
        ws.cell(row=r, column=left_col, value=cohort).font = Font(bold=True)
        ws.cell(row=r, column=left_col + 1, value=base).number_format = (
            "#,##0" if isinstance(base, int) else "#,##0.00"
        )
        for j, col_key in enumerate(columns):
            v = vals.get(col_key)
            cell = ws.cell(row=r, column=left_col + 2 + j, value=v)
            if v is not None:
                cell.number_format = data_fmt
                cell.alignment = Alignment(horizontal="right")

    data_end = data_start + len(rows) - 1
    if data_end >= data_start and color_rule:
        col_letter_start = get_column_letter(left_col + 2)
        col_letter_end = get_column_letter(left_col + 1 + len(columns))
        rng = f"{col_letter_start}{data_start}:{col_letter_end}{data_end}"
        ws.conditional_formatting.add(rng, color_rule)

    return data_end


def _green_to_red(reverse=False):
    """Color scale: low → red, high → green by default. Reverse for churn."""
    if reverse:
        # high churn = bad (red), low = good (green)
        return ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    return ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )


def _signed_churn_scale():
    """For MRR churn (signed): negative = expansion (green), zero = neutral,
    positive = churn (red)."""
    return ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="F8696B",
    )


def write_p9_workbook(out_path: Path, p9: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cohort Analysis"

    cohorts = p9["cohorts"]
    global_max_lt = p9["global_max_lt"]
    lt_cols = list(range(global_max_lt + 1))
    lt_labels = [f"M{lt}" for lt in lt_cols]

    n_base = p9["n_customers_base"]
    mrr_base = p9["cohort_mrr_base"]
    tables = p9["tables"]
    profitable_since = p9["profitable_since"]
    cacs = p9["cacs"]
    gross_margin = p9["gross_margin"]

    row = 1
    title = ws.cell(row=row, column=2, value="Cohort Analysis (Point Nine template)")
    title.font = Font(bold=True, size=18)
    row += 1
    cutoff = p9.get("cutoff_date")
    subtitle_text = (
        f"Gross margin: {gross_margin*100:.0f}%   "
        f"Data through: {cutoff}   "
        f"Cohorts: {len(cohorts)}"
    )
    ws.cell(row=row, column=2, value=subtitle_text).font = Font(italic=True, color="666666")
    row += 2

    # ────────────────────────────────────────────────────────────────
    # Section 1 — Customer Churn
    # ────────────────────────────────────────────────────────────────
    section = ws.cell(row=row, column=2, value="1. Customer Churn")
    section.font = SECTION_FONT
    row += 2

    cust_rows = [
        (c, n_base[c], tables["retained_customers"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "Retained customers in lifetime month",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in cust_rows],
        "#,##0", None,
    )
    row += 2

    churn_rows = [
        (c, n_base[c], tables["churned_customers"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "# of churned customers in lifetime month",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in churn_rows],
        "#,##0", None,
    )
    row += 2

    pct_ret_rows = [
        (c, n_base[c], tables["pct_retained_customers"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% of retained customers in lifetime month",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_ret_rows],
        "0.0%", _green_to_red(),
    )
    row += 2

    pct_churn_base = [
        (c, n_base[c], tables["pct_churned_vs_base_customers"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% of churned customers in lifetime month (relative to base)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_churn_base],
        "0.0%", _green_to_red(reverse=True),
    )
    row += 2

    pct_churn_prev = [
        (c, n_base[c], tables["pct_churned_vs_prev_customers"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% of churned customers in lifetime month (relative to previous month)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_churn_prev],
        "0.0%", _green_to_red(reverse=True),
    )
    row += 3

    # ────────────────────────────────────────────────────────────────
    # Section 2 — MRR Churn
    # ────────────────────────────────────────────────────────────────
    section = ws.cell(row=row, column=2, value="2. MRR Churn")
    section.font = SECTION_FONT
    row += 2

    mrr_ret_rows = [
        (c, mrr_base[c], tables["retained_mrr"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "Retained MRR in lifetime month",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in mrr_ret_rows],
        "$#,##0", None,
    )
    row += 2

    mrr_churn_rows = [
        (c, mrr_base[c], tables["churned_mrr"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "MRR churn in lifetime month (negative = expansion)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in mrr_churn_rows],
        "$#,##0;[Red]-$#,##0", _signed_churn_scale(),
    )
    row += 2

    pct_ret_mrr = [
        (c, mrr_base[c], tables["pct_retained_mrr"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% of retained MRR in lifetime month (NRR)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_ret_mrr],
        "0.0%", _green_to_red(),
    )
    row += 2

    pct_mrr_churn_base = [
        (c, mrr_base[c], tables["pct_churned_vs_base_mrr"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% MRR churn in lifetime month (relative to base)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_mrr_churn_base],
        "0.0%;[Red]-0.0%", _signed_churn_scale(),
    )
    row += 2

    pct_mrr_churn_prev = [
        (c, mrr_base[c], tables["pct_churned_vs_prev_mrr"][c]) for c in cohorts
    ]
    row = _write_table(
        ws, row, 2, "% MRR churn in lifetime month (relative to previous month)",
        lt_labels, [(c, n, {f"M{k}": v for k, v in d.items()}) for c, n, d in pct_mrr_churn_prev],
        "0.0%;[Red]-0.0%", _signed_churn_scale(),
    )
    row += 3

    # ────────────────────────────────────────────────────────────────
    # Section 3 — CAC Payback (only if CACs provided)
    # ────────────────────────────────────────────────────────────────
    if cacs:
        section = ws.cell(row=row, column=2, value="3. CAC Payback")
        section.font = SECTION_FONT
        row += 2

        # Cumulative gross profit table — cohort col, CAC col, then lt_cols
        title_label = (
            f"Cumulated Gross Profit in lifetime month vs. CACs "
            f"(gross margin = {gross_margin*100:.0f}%)"
        )
        _set_header(ws, row, 2, 3 + len(lt_cols) + 1, title_label)
        hdr_row = row + 1
        ws.cell(row=hdr_row, column=2, value="Cohort").font = HEADER_FONT
        ws.cell(row=hdr_row, column=2).fill = HEADER_FILL
        ws.cell(row=hdr_row, column=3, value="CAC").font = HEADER_FONT
        ws.cell(row=hdr_row, column=3).fill = HEADER_FILL
        for j, col in enumerate(lt_labels):
            c = ws.cell(row=hdr_row, column=4 + j, value=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        ws.cell(row=hdr_row, column=4 + len(lt_cols), value="Profitable since").font = HEADER_FONT
        ws.cell(row=hdr_row, column=4 + len(lt_cols)).fill = HEADER_FILL

        data_start = hdr_row + 1
        for i, cohort in enumerate(cohorts):
            r = data_start + i
            cac = cacs.get(cohort, 0)
            ws.cell(row=r, column=2, value=cohort).font = Font(bold=True)
            ws.cell(row=r, column=3, value=cac).number_format = "$#,##0"
            for j, lt in enumerate(lt_cols):
                v = tables["cumulative_gross_profit"][cohort].get(lt)
                cell = ws.cell(row=r, column=4 + j, value=v)
                if v is not None:
                    cell.number_format = "$#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    if cac and v >= cac:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    elif cac:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
            since = profitable_since.get(cohort)
            label = f"M{since}" if since is not None else "Not yet profitable"
            cell = ws.cell(row=r, column=4 + len(lt_cols), value=label)
            cell.font = Font(bold=True, color="2E7D32" if since is not None else "C62828")

        row = data_start + len(cohorts) + 2

    # Column widths — readable
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    for j in range(len(lt_cols)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 12
    ws.column_dimensions[get_column_letter(4 + len(lt_cols))].width = 22

    ws.freeze_panes = "D1"

    wb.save(out_path)

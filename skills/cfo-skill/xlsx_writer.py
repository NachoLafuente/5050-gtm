"""Write a CFO dashboard workbook from the metrics dict produced by cfo.summarize().

Eight sheets: Summary, Cash Flow, Customers, Concentration, AR Aging,
Spend Breakdown, Recurring Vendors, Disclaimer.
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="111111")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEAD_FONT = Font(bold=True, size=10)
GREEN_FILL = PatternFill("solid", fgColor="D6F5D6")
YELLOW_FILL = PatternFill("solid", fgColor="FFF3C4")
RED_FILL = PatternFill("solid", fgColor="FCD9D9")
DISCLAIMER_FILL = PatternFill("solid", fgColor="FFF8DC")


def _money(v: float) -> str:
    if v is None:
        return ""
    return f"{v:,.0f}"


def _pct(v: float) -> str:
    if v is None:
        return ""
    return f"{v:.1f}%"


def _months(v: float) -> str:
    if v is None or math.isinf(v):
        return "∞ (no burn)"
    return f"{v:.1f} mo"


def _flag_runway(months: float) -> PatternFill | None:
    if months == float("inf"):
        return GREEN_FILL
    if months < 12:
        return RED_FILL
    if months < 24:
        return YELLOW_FILL
    return GREEN_FILL


def _flag_concentration(pct: float) -> PatternFill | None:
    if pct > 25:
        return RED_FILL
    if pct > 15:
        return YELLOW_FILL
    return GREEN_FILL


def _flag_dso(dso: float) -> PatternFill | None:
    if dso == 0:
        return None
    if dso < 30:
        return GREEN_FILL
    if dso < 45:
        return None
    if dso < 60:
        return YELLOW_FILL
    return RED_FILL


def _autosize(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 50))
        ws.column_dimensions[letter].width = max_len


def _disclaimer_row(ws, row: int) -> None:
    ws.cell(row, 1, "⚠ Data only — not financial, legal, tax, or investment advice. Verify before acting. Consult a qualified professional.")
    ws.cell(row, 1).fill = DISCLAIMER_FILL
    ws.cell(row, 1).font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


def _header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left")


def _summary_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "CFO Dashboard").font = Font(bold=True, size=18)
    ws.cell(2, 1, f"As of {m['as_of']}").font = Font(italic=True, color="666666")
    _disclaimer_row(ws, 3)

    row = 5
    ws.cell(row, 1, "HEADLINE METRICS").font = SUBHEAD_FONT
    row += 1

    runway_cell_fill = _flag_runway(m["runway_months"])
    headline = [
        ("Total cash balance", _money(m["cash"]["total"]), None),
        ("Trailing 3-month avg net burn", _money(m["burn"]["avg_net_burn"]), None),
        ("Runway", _months(m["runway_months"]), runway_cell_fill),
        ("Burn multiple (net burn / MRR)", f"{m['burn_multiple']:.2f}x" if m["burn_multiple"] else "—", None),
        ("MRR (active customers)", _money(m["mrr"]["mrr"]), None),
        ("ARR", _money(m["mrr"]["arr"]), None),
        ("Active customers", m["mrr"]["active_customers"], None),
        ("Top customer % of MRR", _pct(m["concentration"]["top_1_pct"]), _flag_concentration(m["concentration"]["top_1_pct"])),
        ("Top 5 customers % of MRR", _pct(m["concentration"]["top_5_pct"]), _flag_concentration(m["concentration"]["top_5_pct"])),
        ("AR outstanding (unpaid invoices)", _money(m["ar"]["total"]), None),
        ("DSO (last 90d)", f"{m['dso']:.0f} days", _flag_dso(m["dso"])),
    ]
    for label, val, fill in headline:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, val)
        if fill:
            ws.cell(row, 1).fill = fill
            c.fill = fill
        row += 1

    row += 1
    ws.cell(row, 1, "HEURISTIC FLAGS (not advice)").font = SUBHEAD_FONT
    row += 1
    flags = []
    if m["runway_months"] != float("inf") and m["runway_months"] < 12:
        flags.append("⚠ Runway < 12 months — danger zone heuristic")
    if m["concentration"]["top_1_pct"] > 25:
        flags.append("⚠ Single customer > 25% of MRR — concentration risk")
    elif m["concentration"]["top_1_pct"] > 10:
        flags.append("Note: single customer > 10% of MRR")
    if m["dso"] > 60:
        flags.append("⚠ DSO > 60 days — collection problem")
    if m["burn_multiple"] and m["burn_multiple"] > 2:
        flags.append("⚠ Burn multiple > 2x — efficiency concern")
    if not flags:
        flags = ["No headline heuristics tripped."]
    for f in flags:
        ws.cell(row, 1, f)
        row += 1

    _autosize(ws, 4)


def _cash_flow_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "Cash Flow — Trailing Months").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)
    _header(ws, 4, ["Month", "Inflow", "Outflow", "Net"])
    for i, r in enumerate(m["burn"]["months"], start=5):
        ws.cell(i, 1, r["month"])
        ws.cell(i, 2, r["inflow"])
        ws.cell(i, 3, r["outflow"])
        ws.cell(i, 4, r["net"])
    _autosize(ws, 4)


def _customers_sheet(ws, customers: list[dict]) -> None:
    ws.cell(1, 1, "Customers").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)
    headers = ["customer_id", "customer_name", "status", "signed_up_at", "churned_at", "mrr", "plan"]
    _header(ws, 4, headers)
    for i, c in enumerate(customers, start=5):
        for j, h in enumerate(headers, start=1):
            ws.cell(i, j, c.get(h, ""))
    _autosize(ws, len(headers))


def _concentration_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "Customer Concentration").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)
    ws.cell(4, 1, "Top 1 % of MRR")
    ws.cell(4, 2, _pct(m["concentration"]["top_1_pct"])).fill = _flag_concentration(m["concentration"]["top_1_pct"])
    ws.cell(5, 1, "Top 5 % of MRR")
    ws.cell(5, 2, _pct(m["concentration"]["top_5_pct"])).fill = _flag_concentration(m["concentration"]["top_5_pct"])
    ws.cell(6, 1, "Top 10 % of MRR")
    ws.cell(6, 2, _pct(m["concentration"]["top_10_pct"])).fill = _flag_concentration(m["concentration"]["top_10_pct"])

    _header(ws, 8, ["Customer", "Plan", "MRR", "% of MRR"])
    for i, c in enumerate(m["concentration"]["top_customers"], start=9):
        ws.cell(i, 1, c["name"])
        ws.cell(i, 2, c["plan"])
        ws.cell(i, 3, c["mrr"])
        ws.cell(i, 4, c["pct"] / 100).number_format = "0.0%"
    _autosize(ws, 4)


def _ar_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "AR Aging").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)
    _header(ws, 4, ["Bucket", "Amount"])
    bucket_order = ["current", "1-30", "31-60", "61-90", "90+"]
    for i, b in enumerate(bucket_order, start=5):
        ws.cell(i, 1, b)
        ws.cell(i, 2, m["ar"]["buckets"][b])
        if b in ("61-90", "90+") and m["ar"]["buckets"][b] > 0:
            ws.cell(i, 1).fill = RED_FILL
            ws.cell(i, 2).fill = RED_FILL

    _header(ws, 12, ["Invoice", "Customer", "Due", "Amount", "Days past due", "Bucket"])
    for i, r in enumerate(m["ar"]["rows"], start=13):
        ws.cell(i, 1, r["invoice_id"])
        ws.cell(i, 2, r["customer"])
        ws.cell(i, 3, r["due_at"])
        ws.cell(i, 4, r["amount"])
        ws.cell(i, 5, r["days_past_due"])
        ws.cell(i, 6, r["bucket"])
    _autosize(ws, 6)


def _spend_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "Spend Breakdown").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)

    ws.cell(4, 1, "By Category").font = SUBHEAD_FONT
    _header(ws, 5, ["Category", "Amount", "% of spend"])
    for i, r in enumerate(m["spend_by_category"], start=6):
        ws.cell(i, 1, r["category"])
        ws.cell(i, 2, r["amount"])
        ws.cell(i, 3, r["pct"] / 100).number_format = "0.0%"

    start = 6 + len(m["spend_by_category"]) + 2
    ws.cell(start, 1, "Top Vendors").font = SUBHEAD_FONT
    _header(ws, start + 1, ["Vendor", "Amount", "Transactions", "% of spend"])
    for i, v in enumerate(m["top_vendors"], start=start + 2):
        ws.cell(i, 1, v["vendor"])
        ws.cell(i, 2, v["amount"])
        ws.cell(i, 3, v["count"])
        ws.cell(i, 4, v.get("pct", 0) / 100).number_format = "0.0%"

    start2 = start + 2 + len(m["top_vendors"]) + 2
    ws.cell(start2, 1, "Departmental Burn").font = SUBHEAD_FONT
    _header(ws, start2 + 1, ["Department", "Amount", "% of spend"])
    for i, d in enumerate(m["departmental_burn"], start=start2 + 2):
        ws.cell(i, 1, d["department"])
        ws.cell(i, 2, d["amount"])
        ws.cell(i, 3, d["pct"] / 100).number_format = "0.0%"

    _autosize(ws, 4)


def _recurring_sheet(ws, m: dict) -> None:
    ws.cell(1, 1, "Recurring Vendors (heuristic — verify before cancelling)").font = Font(bold=True, size=14)
    _disclaimer_row(ws, 2)
    _header(ws, 4, ["Vendor", "Occurrences", "Avg amount", "Annualized", "Total observed"])
    for i, r in enumerate(m["recurring_vendors"], start=5):
        ws.cell(i, 1, r["vendor"])
        ws.cell(i, 2, r["occurrences"])
        ws.cell(i, 3, r["avg_amount"])
        ws.cell(i, 4, r["annualized"])
        ws.cell(i, 5, r["total"])
    _autosize(ws, 5)


def _disclaimer_sheet(ws) -> None:
    ws.cell(1, 1, "DISCLAIMER").font = Font(bold=True, size=16, color="C0392B")
    text = [
        "",
        "This workbook displays data pulled from your connected systems.",
        "It is NOT financial, legal, tax, accounting, or investment advice.",
        "",
        "Numbers may be stale, miscategorized, or reflect bugs in upstream systems.",
        "",
        "Heuristic flags (LTV:CAC, runway thresholds, burn multiple, concentration) are",
        "industry rules of thumb. They do not tell you what to do with your business.",
        "",
        "Decisions about hiring, fundraising, distributions, taxes, dividends, solvency,",
        "or any matter with legal or financial consequences must be made by you in",
        "consultation with a qualified CFO, accountant, lawyer, or tax advisor.",
        "",
        "The author and contributors of this software accept no liability for any",
        "decision made based on this workbook's output.",
        "",
        "Always verify critical numbers in the source system before acting on them.",
        "",
        "Generated by /cfo-skill — https://github.com/NachoLafuente/5050-gtm",
    ]
    for i, line in enumerate(text, start=2):
        c = ws.cell(i, 1, line)
        if "DISCLAIMER" in line or "NOT" in line:
            c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 100


def write_workbook(metrics: dict, customers: list[dict], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _summary_sheet(wb.create_sheet("Summary"), metrics)
    _cash_flow_sheet(wb.create_sheet("Cash Flow"), metrics)
    _customers_sheet(wb.create_sheet("Customers"), customers)
    _concentration_sheet(wb.create_sheet("Concentration"), metrics)
    _ar_sheet(wb.create_sheet("AR Aging"), metrics)
    _spend_sheet(wb.create_sheet("Spend"), metrics)
    _recurring_sheet(wb.create_sheet("Recurring Vendors"), metrics)
    _disclaimer_sheet(wb.create_sheet("Disclaimer"))
    wb.save(output_path)

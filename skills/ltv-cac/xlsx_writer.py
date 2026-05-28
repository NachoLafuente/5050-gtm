"""Write the LTV/CAC workbook with conditional formatting.

Three sheets:
  1. Verdict       , inputs, all 4 LTV formulas, payback, ratios, verdict
  2. Sensitivity   , LTV/CAC heatmap across churn × gross margin
  3. Cohort        , 36-month projection of a 100-customer synthetic cohort
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="222222")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=18)
SECTION_FONT = Font(bold=True, size=14)
ITALIC_GRAY = Font(italic=True, color="666666")

VERDICT_FILLS = {
    "red":    PatternFill("solid", fgColor="FFC7CE"),
    "yellow": PatternFill("solid", fgColor="FFEB9C"),
    "green":  PatternFill("solid", fgColor="C6EFCE"),
    "blue":   PatternFill("solid", fgColor="BDD7EE"),
}


def _write_verdict_sheet(ws, out):
    ws.title = "Verdict"
    inputs = out.inputs

    ws.cell(row=1, column=2, value="LTV / CAC Unit Economics").font = TITLE_FONT
    ws.cell(row=2, column=2, value=(
        f"ARPU ${inputs.arpu:,.0f}/mo  ·  Churn {inputs.churn_rate*100:.1f}%  ·  "
        f"Expansion {inputs.expansion_rate*100:.1f}%  ·  GM {inputs.gross_margin*100:.0f}%  ·  "
        f"CAC ${inputs.cac:,.0f}  ·  Inference ${inputs.inference_cost:.0f}/cust/mo"
    )).font = ITALIC_GRAY

    # Inputs block
    ws.cell(row=4, column=2, value="Inputs").font = SECTION_FONT
    inputs_rows = [
        ("ARPU (monthly $)", inputs.arpu, "$#,##0"),
        ("Customer churn rate (monthly)", inputs.churn_rate, "0.00%"),
        ("Net revenue expansion (monthly)", inputs.expansion_rate, "0.00%"),
        ("Gross margin", inputs.gross_margin, "0.00%"),
        ("CAC (per customer $)", inputs.cac, "$#,##0"),
        ("Inference / variable cost ($/cust/mo)", inputs.inference_cost, "$#,##0"),
    ]
    r = 5
    for label, value, fmt in inputs_rows:
        ws.cell(row=r, column=2, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=3, value=value)
        c.number_format = fmt
        r += 1

    # Verdict box
    r += 1
    verdict_cell = ws.cell(row=r, column=2, value=out.verdict)
    verdict_cell.font = Font(bold=True, size=12)
    verdict_cell.fill = VERDICT_FILLS.get(out.verdict_color, PatternFill())
    verdict_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 40

    r += 2

    # NDR
    ws.cell(row=r, column=2, value="NDR (Net Dollar Retention)").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=2, value="Monthly NDR").font = LABEL_FONT
    ws.cell(row=r, column=3, value=out.monthly_ndr).number_format = "0.00%"
    r += 1
    ws.cell(row=r, column=2, value="Annual NDR (compounded)").font = LABEL_FONT
    ws.cell(row=r, column=3, value=out.annual_ndr).number_format = "0.00%"
    r += 2

    # LTV table
    ws.cell(row=r, column=2, value="LTV by formula").font = SECTION_FONT
    r += 1
    headers = ["Formula", "LTV ($)", "LTV / CAC", "Citation"]
    for j, h in enumerate(headers):
        c = ws.cell(row=r, column=2 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r += 1
    data_start = r
    for ltv_res, (_, ratio) in zip(out.ltv_results, out.ltv_cac_ratios):
        ws.cell(row=r, column=2, value=ltv_res.formula).font = LABEL_FONT
        if ltv_res.ltv is None:
            ws.cell(row=r, column=3, value="undefined")
        else:
            ws.cell(row=r, column=3, value=ltv_res.ltv).number_format = "$#,##0"
        if ratio is None:
            ws.cell(row=r, column=4, value="-")
        else:
            cell = ws.cell(row=r, column=4, value=ratio)
            cell.number_format = "0.00"
            if ratio < 1:
                cell.fill = VERDICT_FILLS["red"]
            elif ratio < 3:
                cell.fill = VERDICT_FILLS["yellow"]
            elif ratio <= 5:
                cell.fill = VERDICT_FILLS["green"]
            else:
                cell.fill = VERDICT_FILLS["blue"]
        ws.cell(row=r, column=5, value=ltv_res.citation).font = ITALIC_GRAY
        if ltv_res.notes:
            ws.cell(row=r, column=6, value=ltv_res.notes).font = ITALIC_GRAY
        r += 1
    r += 1

    # CAC Payback
    ws.cell(row=r, column=2, value="CAC Payback").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=2, value="Basic (CAC ÷ gross profit/mo)").font = LABEL_FONT
    if out.cac_payback_basic is not None:
        c = ws.cell(row=r, column=3, value=out.cac_payback_basic)
        c.number_format = "0.0"
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4, value="months")
    else:
        ws.cell(row=r, column=3, value="-")
    r += 1
    ws.cell(row=r, column=2, value="AI-adjusted (subtract inference)").font = LABEL_FONT
    if out.cac_payback_ai_adjusted is not None:
        c = ws.cell(row=r, column=3, value=out.cac_payback_ai_adjusted)
        c.number_format = "0.0"
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4, value="months")
    else:
        ws.cell(row=r, column=3, value="-")
    r += 2

    # Citations
    ws.cell(row=r, column=2, value="Frameworks referenced").font = SECTION_FONT
    r += 1
    refs = [
        "David Skok, “SaaS Metrics 2.0”, for-entrepreneurs.com",
        "a16z, “The 16 Startup Metrics”, a16z.com",
        "Sequoia Capital, “LTV/CAC”, sequoiacap.com",
        "Tomasz Tunguz, “Unit Economics of LLMs”, tomtunguz.com",
    ]
    for ref in refs:
        ws.cell(row=r, column=2, value="• " + ref).font = ITALIC_GRAY
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 60


def _write_sensitivity_sheet(ws, out):
    ws.title = "Sensitivity"
    s = out.sensitivity

    ws.cell(row=1, column=2, value="LTV / CAC Sensitivity").font = TITLE_FONT
    ws.cell(row=2, column=2, value=(
        f"How does LTV/CAC change as churn ({s['row_label']}) and "
        f"gross margin ({s['col_label']}) move? Held constant: "
        f"ARPU ${out.inputs.arpu:,.0f}/mo, CAC ${out.inputs.cac:,.0f}."
    )).font = ITALIC_GRAY

    top = 4
    # Top-left corner label
    corner = ws.cell(row=top, column=2, value=f"{s['row_label']} ↓ / {s['col_label']} →")
    corner.font = HEADER_FONT
    corner.fill = HEADER_FILL
    # Column headers (gross margins)
    for j, col in enumerate(s["cols"]):
        c = ws.cell(row=top, column=3 + j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Rows (churn rates)
    for i_row, row_label in enumerate(s["rows"]):
        r = top + 1 + i_row
        ws.cell(row=r, column=2, value=row_label).font = LABEL_FONT
        for j_col, val in enumerate(s["cells"][i_row]):
            cell = ws.cell(row=r, column=3 + j_col, value=val)
            if val is not None:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right")

    # Color scale: red (low) → yellow → green (high), with manual breakpoints
    n_rows = len(s["rows"])
    n_cols = len(s["cols"])
    if out.inputs.cac > 0:
        rng = (
            f"{get_column_letter(3)}{top+1}:"
            f"{get_column_letter(2 + n_cols)}{top + n_rows}"
        )
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=3, mid_color="FFEB84",
                end_type="num", end_value=8, end_color="63BE7B",
            ),
        )

    # Reference legend
    legend_row = top + n_rows + 2
    ws.cell(row=legend_row, column=2, value="Skok 3:1 reference").font = SECTION_FONT
    legend = [
        ("< 1.0, underwater", "red"),
        ("1.0 – 3.0, tight, careful scaling", "yellow"),
        ("3.0 – 5.0, healthy (Skok target)", "green"),
        ("> 5.0, possibly under-investing", "blue"),
    ]
    for i, (text, color) in enumerate(legend):
        r = legend_row + 1 + i
        cell = ws.cell(row=r, column=2, value=text)
        cell.fill = VERDICT_FILLS[color]
        cell.alignment = Alignment(horizontal="left", indent=1)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for j in range(n_cols):
        ws.column_dimensions[get_column_letter(3 + j)].width = 12


def _write_cohort_sheet(ws, out):
    ws.title = "Cohort Projection"
    cohort = out.cohort

    ws.cell(row=1, column=2, value="Synthetic Cohort Projection").font = TITLE_FONT
    ws.cell(row=2, column=2, value=(
        "Project a 100-customer cohort forward 36 months using your churn, "
        "expansion, and inference inputs. Cumulative gross profit is "
        "compared to total CAC for the cohort to find payback."
    )).font = ITALIC_GRAY

    top = 4
    headers = ["Lifetime month", "Customers retained", "MRR ($)", "Cumulative gross profit ($)", "vs. CAC ($)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=top, column=2 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    cac_total = cohort["cac_total"]
    payback = cohort["payback_month"]
    for i, m in enumerate(cohort["months"]):
        r = top + 1 + i
        ws.cell(row=r, column=2, value=m).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=cohort["customers"][i]).number_format = "#,##0.0"
        ws.cell(row=r, column=4, value=cohort["mrr"][i]).number_format = "$#,##0"
        c = ws.cell(row=r, column=5, value=cohort["cum_gross_profit"][i])
        c.number_format = "$#,##0"
        delta = cohort["cum_gross_profit"][i] - cac_total
        c2 = ws.cell(row=r, column=6, value=delta)
        c2.number_format = "$#,##0;[Red]-$#,##0"
        if cac_total > 0:
            c2.fill = VERDICT_FILLS["green"] if delta >= 0 else VERDICT_FILLS["red"]

    # Payback callout
    callout_row = top + len(cohort["months"]) + 2
    if payback is not None:
        msg = f"Cohort breaks even on CAC at lifetime month {payback} (cumulative GP ≥ ${cac_total:,.0f})."
        ws.cell(row=callout_row, column=2, value=msg).fill = VERDICT_FILLS["green"]
    elif cac_total > 0:
        msg = f"Cohort does not break even within the 36-month horizon (CAC = ${cac_total:,.0f})."
        ws.cell(row=callout_row, column=2, value=msg).fill = VERDICT_FILLS["red"]
    else:
        msg = "No CAC provided, payback section skipped."
        ws.cell(row=callout_row, column=2, value=msg).font = ITALIC_GRAY
    ws.merge_cells(
        start_row=callout_row, start_column=2, end_row=callout_row, end_column=6,
    )
    ws.cell(row=callout_row, column=2).font = Font(bold=True)

    # Add a chart
    chart = LineChart()
    chart.title = "Cumulative gross profit vs CAC"
    chart.x_axis.title = "Lifetime month"
    chart.y_axis.title = "Dollars"
    chart.height = 10
    chart.width = 18
    data = Reference(
        ws,
        min_col=5, max_col=5,
        min_row=top, max_row=top + len(cohort["months"]),
    )
    chart.add_data(data, titles_from_data=True)
    cats = Reference(
        ws,
        min_col=2, max_col=2,
        min_row=top + 1, max_row=top + len(cohort["months"]),
    )
    chart.set_categories(cats)
    ws.add_chart(chart, f"H{top}")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 18


def write_ltv_workbook(out_path: Path, out) -> None:
    wb = Workbook()
    _write_verdict_sheet(wb.active, out)
    _write_sensitivity_sheet(wb.create_sheet("Sensitivity"), out)
    _write_cohort_sheet(wb.create_sheet("Cohort Projection"), out)
    wb.save(out_path)
